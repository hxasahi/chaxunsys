# -*- coding: cp936 -*-
from openpyxl import load_workbook
wb = load_workbook(r'c:\users\administrator\desktop\name.xlsx')
ws = wb.active
def abc(n) :
    nums = ws.max_row
    for i in range(2,nums) :
        if n == ws.cell(row=i,column=1).value :
            print('')
            print('               =================')
            print('               | 姓名：'),
            print(ws.cell(row=i,column=2).value),
            print('  |')
            print('               |               | ')
            print('               |考号：{}  |'.format(ws.cell(row=i,column=1).value))
            print('               |               | ')
            print('               ======成绩=======')
            print('               |语文：{}       |'.format(ws.cell(row=i,column=3).value))
            print('               |数学：{}       |'.format(ws.cell(row=i,column=4).value))
            print('               |英语：{}       |'.format(ws.cell(row=i,column=5).value))
            print('               |物理：{}       |'.format(ws.cell(row=i,column=6).value))
            print('               |化学：{}       |'.format(ws.cell(row=i,column=7).value))
            print('               |生物：{}       |'.format(ws.cell(row=i,column=8).value))
            print('               |总分：{}      |'.format(ws.cell(row=i,column=10).value))
            print('               =================')
n = input('fds')
abc(n)
