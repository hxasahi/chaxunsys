#_*_encoding:cp936_*_
import random
from openpyxl import load_workbook
wb = load_workbook(r'c:\users\administrator\desktop\name.xlsx')
ws = wb.active
print('��ӭʹ�óɼ���ѯ���')
print('*********************')
print('1 �ɼ���ѯ')
print('2 ѧ���༶��ѯ')
print('3 ѧУ�γ̱�')
print('4 ѧУ����')
print('5 ����汾��Ϣ')
print('6 ���ѡȡ')
print('7 �˳�')
print('*********************')
i = input('�밴�¶�Ӧ����ѡ��')
def choice(lists,num):
    i = 1
    while i<=num and num<=len(l):
        print random.choice(lists)
        i = i + 1
nums = ws.max_row
while 1>0 :
    if i == 1 :
        n = input('�����뿼�Ų�ѯ��')
        for i in range(2,nums) :
            if n == ws.cell(row=i,column=1).value :
                print('=================')
                print(' ������'),
                print(ws.cell(row=i,column=2).value)
                print('|���ţ�{}  |'.format(ws.cell(row=i,column=1).value))
                print('======�ɼ�=======')
                print('|���ģ�{}       |'.format(ws.cell(row=i,column=3).value))
                print('|��ѧ��{}       |'.format(ws.cell(row=i,column=4).value))
                print('|Ӣ�{}       |'.format(ws.cell(row=i,column=5).value))
                print('|����{}       |'.format(ws.cell(row=i,column=6).value))
                print('|��ѧ��{}       |'.format(ws.cell(row=i,column=7).value))
                print('|���{}       |'.format(ws.cell(row=i,column=8).value))
                print('|�ܷ֣�{}      |'.format(ws.cell(row=i,column=10).value))
                print('=================')
    elif i == 2 :
        m = input('�����뿼�Ų�ѯ��')
        for i in range(2,nums) :
            if m == ws.cell(row=i,column=1).value :
                print('====ѧ����Ϣ=====')
                print(' ������'),
                print(ws.cell(row=i,column=2).value)
                print('���ţ�{}'.format(ws.cell(row=i,column=1).value))
                print('�༶��'),
                print(ws.cell(row=i,column=11).value)
                print('=================')
    elif i == 3 :
        print('δ���£��޷���ʾ')
    elif i == 4 :
        print('�߶��ӳٿ�ѧ����ѧʱ������֪ͨ��')
    elif i == 5 :
        print('===============================')
        print('�ɼ���ѯ���v4.2')
        print('�������������')
        print('�������ʱ�䣺2020.1.28 19:18')
        print('===============================')
    if i == 6 :
        pass
        '''q = input('������Ҫѡ���������')        
        choice(l,q)'''
    print('**********************')
    print('1 �ɼ���ѯ')
    print('2 �༶��ѯ')
    print('3 ѧУ�γ̱�')
    print('4 ѧУ����')
    print('5 ����汾��Ϣ')
    print('6 ���ѡȡ')
    print('7 �˳�')
    print('**********************') 
    if i == 7 :
        print('****��ӭʹ�óɼ���ѯ���****')
        break
    i = input('�밴�¶�Ӧ����ѡ��') 
      
