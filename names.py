import random
from openpyxl import load_workbook
wb = load_workbook(r'c:\users\administrator\desktop\name.xlsx')
ws = wb.active
l = []
def choice(lists,num):
    i = 1
    while i<=num and num<=len(lists):
        print random.choice(lists)
        i = i + 1
nums = ws.max_row
for p in range(2,nums) :
    names = ws.cell(row = p,column=2).value
    l.append(names)
    p = p + 1
num = input('please enter numbers to choice:')
choice(l,num)
