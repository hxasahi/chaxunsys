#_*_encoding:cp936_*_
import cv2
import numpy as np
import os
import random
from openpyxl import load_workbook
print('')
print('')
print('')
print('')
print('')
print('')
print('')
print('               1 人脸识别'),
print('               2 账号密码')
print('')
xxx=input('               请选择登录方式：') 


while xxx==2:
    print('')
    name=raw_input('               请输入账号名：')
    print('')
    password=raw_input('               请输入密码：')
    break

while xxx==1:
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read('F:/trainer/trainer.yml')
    cascadePath = "C:/Python27/Lib/site-packages/cv2/data/haarcascade_frontalface_default.xml"
    faceCascade = cv2.CascadeClassifier(cascadePath);
 
    font = cv2.FONT_HERSHEY_SIMPLEX
 
    #iniciate id counter
    id = 0
    # Initialize and start realtime video capture
    cam = cv2.VideoCapture(0)
    cam.set(3, 640) # set video widht
    cam.set(4, 480) # set video height
 
    ret, img =cam.read()
    img = cv2.flip(img, 1) # Flip vertically
    gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    # names related to ids: example ==> Marcelo: id=1,  etc
    names = ['None', 'Hx', 'Mom', 'None', 'None', 'None'] 
 

    
    # Define min window size to be recognized as a face
    minW = 0.1*cam.get(3)
    minH = 0.1*cam.get(4)
     
    faces = faceCascade.detectMultiScale( 
        gray,
        scaleFactor = 1.2,
        minNeighbors = 5,
        minSize = (int(minW), int(minH)),
       )
 
    for(x,y,w,h) in faces:
        cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 2)
        id, confidence = recognizer.predict(gray[y:y+h,x:x+w])
 
        # Check if confidence is less them 100 ==> "0" is perfect match 
        if (confidence < 50):
            id = names[id]
            confidence = "  {0}%".format(round(100 - confidence))
        else:
            id = "unknown"
            confidence = "  {0}%".format(round(100 - confidence))
         
        cv2.putText(img, str(id), (x+5,y-5), font, 1, (255,255,255), 2)
        cv2.putText(img, str(confidence), (x+5,y+h-5), font, 1, (255,255,0), 1)  
     
    cv2.imshow('camera',img)
    
 
    k = cv2.waitKey(10) & 0xff # Press 'ESC' for exiting video
    if id == 'Hx' :
        # Do a bit of cleanup
        print("\n [INFO] Exiting Program and cleanup stuff")
        cam.release()
        cv2.destroyAllWindows()

        break
 
    
wb = load_workbook(r'F:\分类\name.xlsx')
ws = wb.active
print('')
print('')
print('                          欢迎使用成绩查询软件')
print('')
print('1 成绩查询        '),
print('2 班级查询      '),
print('3 学校课程表   '),
print('4 学校公告     ')
print('')
print('5 软件版本信息    '),
print('6 随机选取      '),
print('7 退出')
print('')
i = input('               请按下对应数字选择：')
def choice(lists,num):
    i = 1
    while i<=num and num<=len(lists):
        print('               ')
        print random.choice(lists)
        i = i + 1
def kaohao(n) :
    ceshi = 0
    nums = ws.max_row
    for i in range(2,nums+1) :
        if n == ws.cell(row=i,column=1).value :
            print('')
            print('               =================')
            print('               | 姓名：'),
            print(ws.cell(row=i,column=2).value),
            print('  |')
            print('               |               | ')
            print('               |考号：{}  |'.format(ws.cell(row=i,column=1).value))
            print('               |               | ')
            print('               ======成绩=======')
            print('               |语文：{}       |'.format(ws.cell(row=i,column=3).value))
            print('               |数学：{}       |'.format(ws.cell(row=i,column=4).value))
            print('               |英语：{}       |'.format(ws.cell(row=i,column=5).value))
            print('               |物理：{}       |'.format(ws.cell(row=i,column=6).value))
            print('               |化学：{}       |'.format(ws.cell(row=i,column=7).value))
            print('               |生物：{}       |'.format(ws.cell(row=i,column=8).value))
            print('               |总分：{}      |'.format(ws.cell(row=i,column=10).value))
            print('               =================')
            ceshi = 1
        else :
            pass
    if ceshi == 0 :
        print('')
        print('               考号错误！')
        print('')
        n = input('               请重新输入考号查询或按2返回：')
        if n != 2 :
            kaohao(n)
        else :
            pass
        
def xinxi(m) :
    ceshi = 0
    nums = ws.max_row
    for i in range(2,nums+1) :
        if m == ws.cell(row=i,column=1).value :

            print('')
            print('               ====学生信息=====')
            print('               | 姓名：'),
            print(ws.cell(row=i,column=2).value),
            print('  |')
            print('               |               | ')
            print('               |考号：{}  |'.format(ws.cell(row=i,column=1).value))
            print('               |               |')
            print('               |班级：'),
            print(ws.cell(row=i,column=11).value),
            print('  |')
            print('               =================')
            ceshi = 1
        else :
            pass
    if ceshi == 0 :
        print('')
        print('               考号错误！')
        print('')
        n = input('               请重新输入考号查询或按2返回：')
        if n != 2 :
            xinxi(m)
        else :
            pass

nums = ws.max_row

while id == 'Hx' or (name == 'Hx' and password == '0824' ):
    
    if i == 1 :
        print('')
        n = input('               请输入考号查询：')
        kaohao(n)
        
    elif i == 2 :
        print('')
        m = input('               请输入考号查询：')
        xinxi(m)
        
    elif i == 3 :
        print('')
        print('  未更新，无法显示')
        
    elif i == 4 :
        print('')
        print('  高二延迟开学，开学时间另行通知。')
        
    elif i == 5 :
        print('')
        print('               ================================')
        print('               |成绩查询软件v6.1              |')
        print('               |                              |')
        print('               |软件制作：韩旭                |')
        print('               |                              | ')
        print('               |软件制作时间：2020.2.12 16:40 |')
        print('               ================================')
        
    elif i == 6 :
        l = []
        nums = ws.max_row        
        for p in range(2,nums) :
            names = ws.cell(row = p,column=2).value
            l.append(names)
            p = p + 1
        print('')
        q = input('               请输入要选择的人数：')        
        choice(l,q)
        
    elif i == 7 :
        print('')
        print('               ****欢迎使用成绩查询软件****')
        break
    else :
        print('')
        print('               请输入合法字符!')
        
    print('')
    print('               1 继续    '),
    print('2 退出')
    print('')

    c = input('               是否继续？')
    
    while c != 1 and c != 2 :
        print('')
        print('               请输入合法字符!')
        print('')
        print('               1 继续    '),
        print('2 退出')
        print('')
        c = input('               是否继续？')
    if c == 1 :
        print('')
        print('1 成绩查询        '),
        print('2 班级查询      '),
        print('3 学校课程表   '),
        print('4 学校公告     ')
        print('') 
        print('5 软件版本信息    '),
        print('6 随机选取      '),
        print('7 退出')
        print('')     
    elif c == 2 :
        break
    
    i = input('               请按下对应数字选择：')       

