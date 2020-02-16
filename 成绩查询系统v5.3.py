#_*_encoding:cp936_*_
import random
from openpyxl import load_workbook
wb = load_workbook(r'F:\分类\name.xlsx')
ws = wb.active
print('')
print('')
print('                          欢迎使用成绩查询软件')
print('')
print('1 成绩查询        '),
print('2 班级查询      '),
print('3 学校课程表   '),
print('4 学校公告     ')
print('')
print('5 软件版本信息    '),
print('6 随机选取      '),
print('7 退出')
print('')
i = input('               请按下对应数字选择：')
def choice(lists,num):
    i = 1
    while i<=num and num<=len(lists):
        print('               ')
        print random.choice(lists)
        i = i + 1
def kaohao(n) :
    ceshi = 0
    nums = ws.max_row
    for i in range(2,nums+1) :
        if n == ws.cell(row=i,column=1).value :
            print('')
            print('               =================')
            print('               | 姓名：'),
            print(ws.cell(row=i,column=2).value),
            print('  |')
            print('               |               | ')
            print('               |考号：{}  |'.format(ws.cell(row=i,column=1).value))
            print('               |               | ')
            print('               ======成绩=======')
            print('               |语文：{}       |'.format(ws.cell(row=i,column=3).value))
            print('               |数学：{}       |'.format(ws.cell(row=i,column=4).value))
            print('               |英语：{}       |'.format(ws.cell(row=i,column=5).value))
            print('               |物理：{}       |'.format(ws.cell(row=i,column=6).value))
            print('               |化学：{}       |'.format(ws.cell(row=i,column=7).value))
            print('               |生物：{}       |'.format(ws.cell(row=i,column=8).value))
            print('               |总分：{}      |'.format(ws.cell(row=i,column=10).value))
            print('               =================')
            ceshi = 1
        else :
            pass
    if ceshi == 0 :
        print('')
        print('               考号错误！')
        print('')
        n = input('               请重新输入考号查询或按2返回：')
        if n != 2 :
            kaohao(n)
        else :
            pass
        
def xinxi(m) :
    ceshi = 0
    nums = ws.max_row
    for i in range(2,nums+1) :
        if m == ws.cell(row=i,column=1).value :

            print('')
            print('               ====学生信息=====')
            print('               | 姓名：'),
            print(ws.cell(row=i,column=2).value),
            print('  |')
            print('               |               | ')
            print('               |考号：{}  |'.format(ws.cell(row=i,column=1).value))
            print('               |               |')
            print('               |班级：'),
            print(ws.cell(row=i,column=11).value),
            print('  |')
            print('               =================')
            ceshi = 1
        else :
            pass
    if ceshi == 0 :
        print('')
        print('               考号错误！')
        print('')
        n = input('               请重新输入考号查询或按2返回：')
        if n != 2 :
            xinxi(m)
        else :
            pass

nums = ws.max_row

while 1>0 :
    
    if i == 1 :
        print('')
        n = input('               请输入考号查询：')
        kaohao(n)
        
    elif i == 2 :
        print('')
        m = input('               请输入考号查询：')
        xinxi(m)
        
    elif i == 3 :
        print('')
        print('  未更新，无法显示')
        
    elif i == 4 :
        print('')
        print('  高二延迟开学，开学时间另行通知。')
        
    elif i == 5 :
        print('')
        print('               ================================')
        print('               |成绩查询软件v5.3              |')
        print('               |                              |')
        print('               |软件制作：韩旭                |')
        print('               |                              | ')
        print('               |软件制作时间：2020.1.30 12:31 |')
        print('               ================================')
        
    elif i == 6 :
        l = []
        nums = ws.max_row        
        for p in range(2,nums) :
            names = ws.cell(row = p,column=2).value
            l.append(names)
            p = p + 1
        print('')
        q = input('               请输入要选择的人数：')        
        choice(l,q)
        
    elif i == 7 :
        print('')
        print('               ****欢迎使用成绩查询软件****')
        break
    else :
        print('')
        print('               请输入合法字符!')
        
    print('')
    print('               1 继续    '),
    print('2 退出')
    print('')

    c = input('               是否继续？')
    
    while c != 1 and c != 2 :
        print('')
        print('               请输入合法字符!')
        print('')
        print('               1 继续    '),
        print('2 退出')
        print('')
        c = input('               是否继续？')
    if c == 1 :
        print('')
        print('1 成绩查询        '),
        print('2 班级查询      '),
        print('3 学校课程表   '),
        print('4 学校公告     ')
        print('') 
        print('5 软件版本信息    '),
        print('6 随机选取      '),
        print('7 退出')
        print('')     
    elif c == 2 :
        break
    
    i = input('               请按下对应数字选择：')       
