#_*_encoding:cp936_*_
import random
from openpyxl import load_workbook
wb = load_workbook(r'c:\users\administrator\desktop\name.xlsx')
ws = wb.active
print('欢迎使用成绩查询软件')
print('*********************')
print('1 成绩查询')
print('2 学生班级查询')
print('3 学校课程表')
print('4 学校公告')
print('5 软件版本信息')
print('6 随机选取')
print('7 退出')
print('*********************')
i = input('请按下对应数字选择：')
def choice(lists,num):
    i = 1
    while i<=num and num<=len(l):
        print random.choice(lists)
        i = i + 1

while 1>0 :
    if i == 1 :
        n = input('请输入考号查询：')
        num = ws.max_row
        for i in range(2,num) :
            if n == ws.cell(row=i,column=1).value :
                print('=================')
                print(' 姓名：'),
                print(ws.cell(row=i,column=2).value)
                print('|考号：{}  |'.format(ws.cell(row=i,column=1).value))
                print('======成绩=======')
                print('|语文：{}       |'.format(ws.cell(row=i,column=3).value))
                print('|数学：{}       |'.format(ws.cell(row=i,column=4).value))
                print('|英语：{}       |'.format(ws.cell(row=i,column=5).value))
                print('|物理：{}       |'.format(ws.cell(row=i,column=6).value))
                print('|化学：{}       |'.format(ws.cell(row=i,column=7).value))
                print('|生物：{}       |'.format(ws.cell(row=i,column=8).value))
                print('|总分：{}      |'.format(ws.cell(row=i,column=10).value))
                print('=================')
    elif i == 2 :
        m = input('请输入考号查询：')
        pass
        '''if m == p[1] :
            pass
            print('====学生信息=====')
            print('|姓名：{}      |'.format(p[0]))
            print('|考号：{}    |'.format(p[1]))
            print('|班级：{}    |'.format(p[2]))
            print('=================')
        elif m == h[1] :
            print('======成绩=======')
            print('|姓名：{}      |'.format(h[0]))
            print('|考号：{}    |'.format(h[1]))
            print('|班级：{}    |'.format(h[2]))
            print('=================')'''
    elif i == 3 :
        print('未更新，无法显示')
    elif i == 4 :
        print('高二延迟开学，开学时间另行通知。')
    elif i == 5 :
        print('===============================')
        print('成绩查询软件v4.1')
        print('软件制作：韩旭')
        print('软件制作时间：2020.1.28 19:18')
        print('===============================')
    if i == 6 :
        pass
        '''q = input('请输入要选择的人数：')        
        choice(l,q)'''
    print('**********************')
    print('1 成绩查询')
    print('2 班级查询')
    print('3 学校课程表')
    print('4 学校公告')
    print('5 软件版本信息')
    print('6 随机选取')
    print('7 退出')
    print('**********************')
    i = input('请按下对应数字选择：')  
    if i == 7 :
        print('****欢迎使用成绩查询软件****')
        break
      
